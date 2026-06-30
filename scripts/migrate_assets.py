"""One-off migration: build `mt_asset_list` from the revised Excel asset register
(both W-202 and A-185 sheets), then drop the old `mt_machine_list` table.

Credentials are read from app/.env (DATABASE_URL) — never hardcoded or passed
on the command line.

Usage:
    python scripts/migrate_assets.py --dry-run   # parse + validate only, no DB
    python scripts/migrate_assets.py             # full migration (writes prod)
"""
import argparse
import os
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

BACKEND_DIR = Path(__file__).resolve().parents[1]
ENV_PATH = BACKEND_DIR / "app" / ".env"
XLSX_PATH = BACKEND_DIR.parent / "Asset Full revised A185 & W202.xlsx"

NEW_TABLE = "mt_asset_list"
OLD_TABLE = "mt_machine_list"

# (column_name, max_len or None for non-string / unlimited)
COLUMN_LIMITS = {
    "asset_id": 32,
    "building": 16,
    "asset_name": 255,
    "category": 64,
    "sub_category": 64,
    "sub_location": 255,
    "model_no": 255,
    "serial_no": 255,
    "power_load": 128,
    "condition": 32,
    "assigned_to": 128,
    "warranty_amc_expiry": 64,
}

DDL = f"""
DROP TABLE IF EXISTS {NEW_TABLE};
CREATE TABLE {NEW_TABLE} (
    id                  SERIAL PRIMARY KEY,
    asset_id            VARCHAR(32) UNIQUE,
    building            VARCHAR(16)  NOT NULL,
    asset_name          VARCHAR(255) NOT NULL,
    category            VARCHAR(64),
    sub_category        VARCHAR(64),
    sub_location        VARCHAR(255),
    quantity            INTEGER,
    revised_count_2026  INTEGER,
    model_no            VARCHAR(255),
    serial_no           VARCHAR(255),
    power_load          VARCHAR(128),
    purchase_date       DATE,
    purchase_value      NUMERIC(14,2),
    condition           VARCHAR(32),
    assigned_to         VARCHAR(128),
    warranty_amc_expiry VARCHAR(64),
    remarks             TEXT
);
CREATE INDEX ix_{NEW_TABLE}_building     ON {NEW_TABLE} (building);
CREATE INDEX ix_{NEW_TABLE}_sub_category ON {NEW_TABLE} (sub_category);
"""

INSERT_SQL = text(f"""
INSERT INTO {NEW_TABLE}
    (asset_id, building, asset_name, category, sub_category, sub_location,
     quantity, revised_count_2026, model_no, serial_no, power_load,
     purchase_date, purchase_value, condition, assigned_to,
     warranty_amc_expiry, remarks)
VALUES
    (:asset_id, :building, :asset_name, :category, :sub_category, :sub_location,
     :quantity, :revised_count_2026, :model_no, :serial_no, :power_load,
     :purchase_date, :purchase_value, :condition, :assigned_to,
     :warranty_amc_expiry, :remarks)
""")


def s(v):
    """Clean a string cell -> stripped str or None."""
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def as_int(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def as_num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None  # free-text / blank purchase dates are dropped


def build_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    rows = []
    counters = {"W-202": 0, "A-185": 0}

    def next_id(building):
        counters[building] += 1
        prefix = "W202" if building == "W-202" else "A185"
        return f"{prefix}-{counters[building]:04d}"

    # ---- W-202 sheet: header row 2, data from row 3 ----
    ws = wb["W-202 Asset List Revised"]
    for r in ws.iter_rows(min_row=3, values_only=True):
        name = s(r[1])
        if not name:
            continue  # spacer / empty row
        rows.append({
            "asset_id": next_id("W-202"),
            "building": "W-202",
            "asset_name": name,
            "category": s(r[4]),
            "sub_category": s(r[5]),
            "sub_location": s(r[7]),
            "quantity": as_int(r[2]),
            "revised_count_2026": as_int(r[3]),
            "model_no": None,
            "serial_no": None,
            "power_load": None,
            "purchase_date": as_date(r[8]),
            "purchase_value": as_num(r[9]),
            "condition": s(r[10]),
            "assigned_to": s(r[11]),
            "warranty_amc_expiry": s(r[12]),
            "remarks": s(r[13]),
        })

    # ---- A-185 sheet: header row 2, data from row 3 ----
    ws = wb["A-185 Asset revised"]
    for r in ws.iter_rows(min_row=3, values_only=True):
        name = s(r[1])
        if not name:
            continue
        rows.append({
            "asset_id": next_id("A-185"),
            "building": "A-185",
            "asset_name": name,
            "category": s(r[2]),
            "sub_category": s(r[3]),
            "sub_location": s(r[5]),
            "quantity": as_int(r[6]),
            "revised_count_2026": as_int(r[7]),
            "model_no": s(r[8]),
            "serial_no": s(r[9]),
            "power_load": s(r[10]),
            "purchase_date": None,
            "purchase_value": None,
            "condition": s(r[11]),
            "assigned_to": s(r[12]),
            "warranty_amc_expiry": s(r[13]),
            "remarks": s(r[14]),
        })

    wb.close()
    return rows, counters


def validate(rows):
    """Abort (no truncation) if any string value exceeds its column size."""
    problems = []
    for i, row in enumerate(rows):
        for col, limit in COLUMN_LIMITS.items():
            val = row.get(col)
            if isinstance(val, str) and len(val) > limit:
                problems.append(f"  row {i} {col!r}={val!r} len={len(val)} > {limit}")
    if problems:
        print("VALIDATION FAILED -- values exceed column sizes:")
        print("\n".join(problems))
        sys.exit(1)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true",
                    help="parse + validate only; do not touch the database")
    args = ap.parse_args()

    if not XLSX_PATH.exists():
        sys.exit(f"Excel not found: {XLSX_PATH}")

    rows, counters = build_rows()
    validate(rows)

    print(f"Parsed rows: W-202={counters['W-202']}, A-185={counters['A-185']}, "
          f"total={len(rows)}")
    print("Sample (first row):", rows[0])
    print("Sample (first A-185):", next(r for r in rows if r['building'] == 'A-185'))

    if args.dry_run:
        print("\n[dry-run] no database changes made.")
        return

    load_dotenv(ENV_PATH)
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        sys.exit(f"DATABASE_URL not found in {ENV_PATH}")

    engine = create_engine(db_url, future=True)
    with engine.begin() as conn:
        for stmt in DDL.strip().split(";"):
            if stmt.strip():
                conn.execute(text(stmt))
        conn.execute(INSERT_SQL, rows)

        inserted = conn.execute(text(f"SELECT COUNT(*) FROM {NEW_TABLE}")).scalar()
        expected = len(rows)
        if inserted != expected:
            raise RuntimeError(
                f"Row count mismatch: inserted={inserted}, expected={expected} "
                f"-- rolling back, {OLD_TABLE} NOT dropped.")
        print(f"Inserted {inserted} rows into {NEW_TABLE} (verified).")

        conn.execute(text(f"DROP TABLE IF EXISTS {OLD_TABLE}"))
        print(f"Dropped {OLD_TABLE}.")

    print("Migration complete.")


if __name__ == "__main__":
    main()

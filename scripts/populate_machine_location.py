"""Add + populate machines.building and machines.sub_location from
W202_and_A185_Asset_List.xlsx, matched by machines.id == machine_id (214/214).
Targets the SQLite DB the backend uses (factoryops.db).

    python -m scripts.populate_machine_location --dry-run
    python -m scripts.populate_machine_location
"""
import argparse
from collections import Counter
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text

BACKEND = Path(__file__).resolve().parents[1]
XLSX = BACKEND / "W202_and_A185_Asset_List.xlsx"
DB_URL = "sqlite:///./factoryops.db"


def build_map():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    m = {}
    for sn in wb.sheetnames:
        building = "W-202" if sn.startswith("W202") else "A-185"
        for r in wb[sn].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                m[str(r[0]).strip()] = {
                    "building": building,
                    "sub_location": str(r[1]).strip() if r[1] else None,
                }
    wb.close()
    return m


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    mp = build_map()
    eng = create_engine(DB_URL, future=True)
    with eng.begin() as c:
        cols = [r[1] for r in c.execute(text("PRAGMA table_info(machines)")).all()]
        rows = c.execute(text("SELECT id FROM machines")).all()

        ups, miss = [], []
        for r in rows:
            v = mp.get(r.id)
            if v:
                ups.append({"id": r.id, **v})
            else:
                miss.append(r.id)

        print(f"machines={len(rows)} matched={len(ups)} miss={len(miss)} "
              f"has_building={'building' in cols} has_sub_location={'sub_location' in cols}")
        if miss:
            print("miss:", miss[:10])
        print("building dist:", dict(Counter(u["building"] for u in ups)))

        if args.dry_run:
            print("[dry-run] no changes.")
            return

        if "building" not in cols:
            c.execute(text("ALTER TABLE machines ADD COLUMN building VARCHAR(16)"))
            print("added column machines.building")
        if "sub_location" not in cols:
            c.execute(text("ALTER TABLE machines ADD COLUMN sub_location VARCHAR(255)"))
            print("added column machines.sub_location")
        c.execute(text(
            "UPDATE machines SET building=:building, sub_location=:sub_location WHERE id=:id"), ups)
        filled = c.execute(text("SELECT COUNT(*) FROM machines WHERE building IS NOT NULL")).scalar()
        print(f"APPLIED. building filled: {filled}/{len(rows)}")


if __name__ == "__main__":
    main()

"""Seed the FactoryOps DB from the two source Excel files.

Run with:
    python -m scripts.seed
"""
from __future__ import annotations

import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

# Ensure project root is on sys.path when run as a script
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.database import LocalBase, local_engine, SessionLocal  # noqa: E402
from app.models import (  # noqa: E402
    Plant, Floor, Machine, User, UserMachineAssignment,
    FloorUtilityReading,
)
from app.auth import hash_password  # noqa: E402
from app.utils import parse_kw, parse_qty, infer_machine_type, categorise  # noqa: E402


MACHINE_XLSX = Path(r"d:\Maintenance module\FactoryOps\machine-list.xlsx")
FLOOR_XLSX = Path(r"d:\Maintenance module\Floorwise utility dada.xlsx")

PLANT_ID = "plant-1"
PLANT_NAME = "Candor Foods — Main Plant"


def _floor_id(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.strip().lower()).strip("-")
    return f"floor-{slug}"


# Canonical floor names — merges casing/spelling variants from the two source
# workbooks so the same floor in machine-list and utility data line up.
_FLOOR_CANON: dict[str, str] = {
    "lower basement": "Lower basement",
    "upper basement": "Upper basement",
    "upper basemant": "Upper basement",  # typo in machine list
    "service floor": "Service floor",
    "ground floor": "Ground floor",
    "1st floor": "1st floor",
    "1st floor mazzenine": "1st floor mezzanine",
    "1st floor mezzanine": "1st floor mezzanine",
    "2nd floor": "2nd floor",
    "2nd floor mazenine": "2nd floor mezzanine",
    "2nd floor mezzanine": "2nd floor mezzanine",
    "office": "Office",
    "terrace": "Terrace",
    "tarrace": "Terrace",  # typo in utility sheet
    "ac": "AC",
    "old lift": "Old lift",
}


def _canon_floor(name: str) -> str:
    key = (name or "").strip().lower()
    return _FLOOR_CANON.get(key, (name or "Unassigned").strip())


def seed_plant_and_users(db):
    if db.get(Plant, PLANT_ID) is None:
        db.add(Plant(id=PLANT_ID, name=PLANT_NAME))

    users = [
        ("op-001", "operator1", "Operator One", "pass123", "OPERATOR"),
        ("op-002", "operator2", "Operator Two", "pass123", "OPERATOR"),
        ("tech-001", "technician1", "Technician One", "pass123", "TECHNICIAN"),
    ]
    for uid, uname, name, pw, role in users:
        if db.get(User, uid) is None:
            db.add(User(
                id=uid,
                username=uname,
                name=name,
                password_hash=hash_password(pw),
                role=role,
                plant_id=PLANT_ID,
            ))
    db.commit()


def seed_machines(db) -> dict[str, str]:
    """Returns {floor_name_normalized: floor_id} for the floor-utility step."""
    wb = openpyxl.load_workbook(MACHINE_XLSX, data_only=True)
    ws = wb["Sheet1"]

    rows = list(ws.iter_rows(values_only=True))
    # Header is row index 1: (None, 'machine_id', 'floor', 'machine_name', ...)
    data_rows = rows[2:]

    floor_map: dict[str, str] = {}

    for r in data_rows:
        if not r or len(r) < 10:
            continue
        _, machine_id, floor, machine_name, company, model_no, serial_no, rated_kw, rated_amps, quantity = r[:10]
        if not machine_id:
            continue

        # Floor (canonicalised so casing/spelling variants merge)
        floor_name = _canon_floor(floor or "Unassigned")
        fid = _floor_id(floor_name)
        if fid not in floor_map:
            existing = db.get(Floor, fid)
            if existing is None:
                db.add(Floor(id=fid, plant_id=PLANT_ID, name=floor_name))
            floor_map[fid] = floor_name

        existing_m = db.get(Machine, machine_id)
        if existing_m is not None:
            continue

        machine = Machine(
            id=machine_id,
            code=machine_id,
            name=(machine_name or "Unnamed").strip(),
            location=floor_name,
            plant_id=PLANT_ID,
            floor_id=fid,
            rated_kw=parse_kw(rated_kw),
            load_factor=0.7,
            load_factor_source="ASSUMED",
            criticality="C",
            expected_run_hours=8.0,
            current_status="IDLE",
            machine_type=infer_machine_type(machine_name or ""),
            company=(str(company).strip() if company and str(company).strip() not in ("`",) else None),
            model_no=(str(model_no).strip() if model_no else None),
            serial_no=(str(serial_no).strip() if serial_no else None),
            quantity=parse_qty(quantity),
            rated_amps=(str(rated_amps).strip() if rated_amps else None),
            updated_at=datetime.utcnow(),
        )
        db.add(machine)
    db.commit()

    # Build a name→id map for the utility seed
    name_to_id: dict[str, str] = {}
    for f in db.query(Floor).all():
        name_to_id[f.name.strip().lower()] = f.id
    return name_to_id


# Map column header in the utility sheet -> floor name we use in machine table.
# (Spelling variants in the source workbook get normalised here.)
UTILITY_COL_TO_FLOOR = {
    "Lower basement": "Lower basement",
    "Upper basement": "Upper basement",
    "service floor": "Service floor",
    "Ground floor": "Ground floor",
    "1st floor": "1st floor",
    "2nd floor": "2nd floor",
    "Office": "Office",
    "Tarrace": "Terrace",      # source typo
    "AC": "AC",                 # plant-wide AC sub-meter (no machine list)
    "Old lift": "Old lift",
}


def seed_floor_utility(db, name_to_id: dict[str, str]):
    wb = openpyxl.load_workbook(FLOOR_XLSX, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return

    # Row 0 holds floor names at even columns, "KWH" header at odd columns on row 1.
    header = rows[0]
    # Build column index map: (meter_col, kwh_col, floor_id)
    cols: list[tuple[int, int, str]] = []
    for idx, val in enumerate(header):
        if not isinstance(val, str):
            continue
        floor_label = val.strip()
        if floor_label not in UTILITY_COL_TO_FLOOR:
            continue
        target_name = UTILITY_COL_TO_FLOOR[floor_label]
        fid = name_to_id.get(target_name.strip().lower())
        if fid is None:
            # Create a synthetic floor for AC / Old lift which don't appear in
            # the machine list.
            fid = _floor_id(target_name)
            if db.get(Floor, fid) is None:
                db.add(Floor(id=fid, plant_id=PLANT_ID, name=target_name))
                db.commit()
                name_to_id[target_name.strip().lower()] = fid
        cols.append((idx, idx + 1, fid))

    # Wipe existing rows so re-runs are idempotent.
    db.query(FloorUtilityReading).delete()
    db.commit()

    seen: set[tuple[str, date]] = set()

    for r_idx in range(2, len(rows)):
        row = rows[r_idx]
        if not row:
            continue
        raw_date = row[0]
        if raw_date is None:
            continue
        d: date | None = None
        if isinstance(raw_date, datetime):
            d = raw_date.date()
        elif isinstance(raw_date, str):
            s = raw_date.strip()
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
                try:
                    d = datetime.strptime(s, fmt).date()
                    break
                except ValueError:
                    continue
        if d is None:
            continue

        for meter_col, kwh_col, fid in cols:
            meter = row[meter_col] if meter_col < len(row) else None
            kwh = row[kwh_col] if kwh_col < len(row) else None
            if meter is None and kwh is None:
                continue
            try:
                meter_v = float(meter) if meter is not None else None
            except (TypeError, ValueError):
                meter_v = None
            try:
                kwh_v = float(kwh) if kwh is not None else None
            except (TypeError, ValueError):
                kwh_v = None
            key = (fid, d)
            if key in seen:
                continue
            seen.add(key)
            db.add(FloorUtilityReading(
                floor_id=fid,
                reading_date=d,
                meter_reading=meter_v,
                daily_kwh=kwh_v,
            ))
    db.commit()


def main():
    LocalBase.metadata.create_all(bind=local_engine)
    db = SessionLocal()
    try:
        print("[seed] plant + users ...")
        seed_plant_and_users(db)
        print("[seed] machines from", MACHINE_XLSX)
        name_to_id = seed_machines(db)
        print(f"  floors created: {len(name_to_id)}")
        print("[seed] floor utility readings from", FLOOR_XLSX)
        seed_floor_utility(db, name_to_id)
        m_count = db.query(Machine).count()
        f_count = db.query(Floor).count()
        u_count = db.query(FloorUtilityReading).count()
        print(f"[seed] done -- machines={m_count}, floors={f_count}, "
              f"utility_rows={u_count}")
    finally:
        db.close()


if __name__ == "__main__":
    main()

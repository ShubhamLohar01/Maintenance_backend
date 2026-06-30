"""Add + populate machines.category from W202_and_A185_Asset_List.xlsx.

Match key: machines.id == Excel machine_id (verified 214/214). Targets the
SQLite DB the local backend actually uses (factoryops.db).

    python -m scripts.populate_machine_categories --dry-run   # preview
    python -m scripts.populate_machine_categories             # apply
"""
import argparse
from collections import Counter
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text

BACKEND = Path(__file__).resolve().parents[1]
XLSX = BACKEND / "W202_and_A185_Asset_List.xlsx"
DB_URL = "sqlite:///./factoryops.db"


def build_map():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    m = {}
    for sn in wb.sheetnames:
        for r in wb[sn].iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            cat = r[10] if len(r) > 10 else None
            if cat:
                m[str(r[0]).strip()] = str(cat).strip()
    wb.close()
    return m


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    cat_map = build_map()
    eng = create_engine(DB_URL, future=True)
    with eng.begin() as c:
        cols = [r[1] for r in c.execute(text("PRAGMA table_info(machines)")).all()]
        has_col = "category" in cols
        rows = c.execute(text("SELECT id FROM machines")).all()

        updates, misses = [], []
        for r in rows:
            cat = cat_map.get(r.id)
            if cat:
                updates.append({"id": r.id, "cat": cat})
            else:
                misses.append(r.id)

        print(f"machines={len(rows)} matched={len(updates)} misses={len(misses)} "
              f"category_col_exists={has_col}")
        if misses:
            print("misses:", misses[:10])
        dist = Counter(u["cat"] for u in updates)
        print("distribution:", dict(sorted(dist.items(), key=lambda x: -x[1])))

        if args.dry_run:
            print("[dry-run] no changes.")
            return

        if not has_col:
            c.execute(text("ALTER TABLE machines ADD COLUMN category VARCHAR(64)"))
            print("added column machines.category")
        c.execute(text("UPDATE machines SET category = :cat WHERE id = :id"), updates)
        filled = c.execute(text("SELECT COUNT(*) FROM machines WHERE category IS NOT NULL")).scalar()
        print(f"APPLIED. machines with category: {filled}/{len(rows)}")


if __name__ == "__main__":
    main()

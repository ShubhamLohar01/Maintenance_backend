"""One-off: set mt_asset_list.category from W202_and_A185_Asset_List.xlsx.

Each existing RDS row is matched to the Excel on (building, asset_name, floor)
where floor = sub_location. Pre-verified: all 864 rows match exactly one Excel
category (0 conflicts, 0 misses). RDS creds come from app/.env (DATABASE_URL).

    python -m scripts.update_asset_categories --dry-run   # preview only
    python -m scripts.update_asset_categories             # apply
"""
import argparse
import os
import re
from collections import Counter
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

BACKEND = Path(__file__).resolve().parents[1]
XLSX = BACKEND / "W202_and_A185_Asset_List.xlsx"


def norm(v):
    return re.sub(r"\s+", " ", str(v).strip().lower()) if v not in (None, "") else ""


def build_map():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    m = {}
    for sn in wb.sheetnames:
        building = "W-202" if sn.startswith("W202") else "A-185"
        for r in wb[sn].iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            cat = r[10] if len(r) > 10 else None
            if not cat:
                continue
            m[(building, norm(r[2]), norm(r[1]))] = str(cat).strip()
    wb.close()
    return m


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    load_dotenv(BACKEND / "app" / ".env")
    eng = create_engine(os.environ["DATABASE_URL"], future=True)
    cat_map = build_map()

    with eng.begin() as c:
        rows = c.execute(text(
            "SELECT id, building, asset_name, sub_location FROM mt_asset_list")).all()
        updates, misses = [], []
        for row in rows:
            cat = cat_map.get((row.building, norm(row.asset_name), norm(row.sub_location)))
            if cat is None:
                misses.append((row.asset_name, row.sub_location))
            else:
                updates.append({"id": row.id, "cat": cat})

        print(f"rows={len(rows)}  matched={len(updates)}  misses={len(misses)}")
        if misses:
            print("ABORT: unmatched rows, nothing written. examples:", misses[:8])
            return

        dist = Counter(u["cat"] for u in updates)
        print("category distribution to set:")
        for k, v in sorted(dist.items(), key=lambda x: -x[1]):
            print(f"   {k!r}: {v}")

        if args.dry_run:
            print("[dry-run] no changes made.")
            return

        c.execute(text("UPDATE mt_asset_list SET category = :cat WHERE id = :id"), updates)
        remaining = c.execute(text(
            "SELECT COUNT(*) FROM mt_asset_list WHERE category = 'Machinery'")).scalar()
        print(f"\nAPPLIED. rows still 'Machinery': {remaining}")


if __name__ == "__main__":
    main()
